# -*- coding: utf-8 -*-

import asyncio
import sys
import re
import os
import pandas as pd
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse

sys.stdout.reconfigure(encoding='utf-8')
MAX_SCROLL_ATTEMPTS = 6
OUTPUT_DIR = "." 


def parse_rating_and_reviews(text):
    rating = "N/A"
    reviews = "N/A"
    if not text:
        return rating, reviews
    
    text = text.strip().replace(",", "")
    match = re.search(r"([\d.]+)\s*\(([\d]+)\)", text)
    if not match:
        match = re.search(r"([\d.]+)\s*stars?\s*([\d]+)", text, re.IGNORECASE)
        
    if match:
        rating = match.group(1)
        reviews = match.group(2)
    return rating, reviews


def clean_phone(raw):
    digits = re.sub(r'[^\d+]', '', raw)
    if digits.startswith('971'):
        return '+' + digits
    elif digits.startswith('0'):
        return '+971' + digits[1:]
    elif digits.startswith('+'):
        return digits
    else:
        return '+971' + digits
    
def short_url(url):
    if not url or url == "N/A":
        return "N/A"
    try:
        return urlparse(url).netloc.replace("www.", "")
    except:
        return url

def save_excel(results, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dubai Medical Leads"

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['#', 'Business Name', 'Phone', 'Rating', 'Reviews', 'Maps URL', 'Website']
    col_widths = [5, 40, 20, 10, 10, 25, 30]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30
# -*- coding: utf-8 -*-

import asyncio
import sys
import re
import os
import pandas as pd
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse

sys.stdout.reconfigure(encoding='utf-8')
MAX_SCROLL_ATTEMPTS = 6
OUTPUT_DIR = "." 

def parse_rating_and_reviews(text):
    rating = "N/A"
    reviews = "N/A"
    if not text:
        return rating, reviews
    
    text = text.strip().replace(",", "")
    # دعم اللغة العربية
    r_ar1 = re.search(r"([\d.]+)\s*نجمة", text)
    r_ar2 = re.search(r"([\d,]+)\s*مراجعة", text)
    if r_ar1: rating = r_ar1.group(1)
    if r_ar2: reviews = r_ar2.group(1)
    if r_ar1 or r_ar2: return rating, reviews

    # دعم اللغة الإنجليزية
    match = re.search(r"([\d.]+)\s*\(([\d]+)\)", text)
    if not match:
        match = re.search(r"([\d.]+)\s*stars?\s*([\d]+)", text, re.IGNORECASE)
        
    if match:
        rating = match.group(1)
        reviews = match.group(2)
    return rating, reviews

def clean_phone(raw):
    if not raw or raw == "N/A": return "N/A"
    digits = re.sub(r'[^\d+]', '', raw)
    if digits.startswith('971'): return '+' + digits
    elif digits.startswith('00971'): return '+' + digits[2:]
    elif digits.startswith('0'): return '+971' + digits[1:]
    elif digits.startswith('+'): return digits
    else: return '+971' + digits
    
def short_url(url):
    if not url or url == "N/A": return "N/A"
    try: return urlparse(url).netloc.replace("www.", "")
    except: return url

def save_excel(results, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dubai Medical Leads"

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['#', 'Business Name', 'Phone', 'Rating', 'Reviews', 'Maps URL', 'Website']
    col_widths = [5, 40, 20, 10, 10, 25, 30]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30
    link_font = Font(name='Arial', color='0563C1', underline='single', size=10)
    normal_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    for row_idx, item in enumerate(results, 2):
        row_fill = PatternFill('solid', start_color='EBF3FB') if row_idx % 2 == 0 else PatternFill('solid', start_color='FFFFFF')

        cols_data = [
            (1, row_idx - 1, center_align, normal_font, False),
            (2, item['Business Name'], left_align, normal_font, False),
            (3, item['Phone'], center_align, normal_font, False),
            (4, item['Rating'], center_align, normal_font, False),
            (5, item['Reviews'], center_align, normal_font, False),
            (6, 'Open Map' if item['Maps URL'] != 'N/A' else 'N/A', center_align, link_font if item['Maps URL'] != 'N/A' else normal_font, True),
            (7, short_url(item['Website']) if item['Website'] != 'N/A' else 'N/A', center_align, link_font if item['Website'] != 'N/A' else normal_font, True)
        ]

        for col_num, val, align, font, is_link in cols_data:
            c = ws.cell(row=row_idx, column=col_num, value=val)
            c.font = font
            c.alignment = align
            c.fill = row_fill
            c.border = thin_border
            if is_link and val not in ['N/A']:
                c.hyperlink = item['Maps URL'] if col_num == 6 else item['Website']

        ws.row_dimensions[row_idx].height = 40

    ws.freeze_panes = 'A2'
    wb.save(filepath)

async def scrape_google_maps(search_query, total_results_needed=80):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800}
        )
        page = await context.new_page()

        print(f"[*] Searching for: {search_query}")
        query = search_query.replace(" ", "+")
        
        await page.goto(f"https://www.google.com/maps/search/{query}")
        await page.wait_for_timeout(5000)

        results = []
        seen = set()
        scroll_attempts = 0
        last_cards_count = 0

        print("[*] Progress: Extracting data (Dual-Layer Strategy)...")

        while len(results) < total_results_needed:
            if scroll_attempts >= MAX_SCROLL_ATTEMPTS:
                print("[!] Max scroll reached")
                break

            cards = await page.query_selector_all('div[role="article"]')
            new_cards_count = len(cards) - last_cards_count
            print(f"Cards found: {len(cards)} (+{new_cards_count} new)")

            if new_cards_count == 0 and len(cards) > 0:
                scroll_attempts += 1
            else:
                scroll_attempts = 0

            for card in cards[last_cards_count:]:
                if len(results) >= total_results_needed: break
                try:
                    # 1. استخراج العنوان (طريقتان كما طلبت)
                    name = "N/A"
                    link_el = await card.query_selector("a.hfpxzc")
                    if link_el:
                        name = await link_el.get_attribute("aria-label")
                    
                    if not name or name == "N/A":
                        title_el = await card.query_selector("div.qBF1Pd")
                        if title_el: name = await title_el.inner_text()

                    if not name or name == "N/A" or name in seen:
                        continue

                    url = await link_el.get_attribute("href") if link_el else "N/A"

                    # 2. استخراج التقييمات والمراجعات (طريقتان كما طلبت)
                    rating, reviews = "N/A", "N/A"
                    # الطريقة الأولى (aria-label)
                    stars_el = await card.query_selector('span.ZkP5Je, span[role="img"][aria-label*="star"], span[role="img"][aria-label*="نجمة"]')
                    if stars_el:
                        raw_text = await stars_el.get_attribute('aria-label')
                        if raw_text: rating, reviews = parse_rating_and_reviews(raw_text)
                    
                    # الطريقة الثانية (الكلاسات المباشرة)
                    if rating == "N/A":
                        r_el = await card.query_selector('span.MW4etd')
                        rev_el = await card.query_selector('span.UY7F9')
                        if r_el: rating = await r_el.inner_text()
                        if rev_el: reviews = (await rev_el.inner_text()).strip('()')

                    # 3. استخراج الهاتف والموقع من الكرت الخارجي مباشرة (سريع)
                    phone, website = "N/A", "N/A"
                    
                    phone_el = await card.query_selector('span.UsdlK')
                    if phone_el: phone = clean_phone(await phone_el.inner_text())

                    web_el = await card.query_selector('a[data-value="Website"], a.lcr4fd')
                    if web_el: website = await web_el.get_attribute('href')

                    # 4. الطبقة العميقة: إذا أخفت واجهة الصور الهاتف أو الموقع، ندخل للوحة الجانبية
                    if (phone == "N/A" or website == "N/A") and link_el:
                        await link_el.click()
                        
                        matched = False
                        for _ in range(8):
                            panel_title_el = await page.query_selector('h1.DUwDvf')
                            if panel_title_el:
                                panel_title = await panel_title_el.inner_text()
                                if name.strip() in panel_title.strip() or panel_title.strip() in name.strip():
                                    matched = True
                                    break
                            await page.wait_for_timeout(200)

                        if matched:
                            if phone == "N/A":
                                phone_button = await page.query_selector('button[data-item-id^="phone:tel:"]')
                                if phone_button:
                                    phone_raw = await phone_button.get_attribute('data-item-id')
                                    phone = clean_phone(phone_raw.replace("phone:tel:", "").strip())
                                else:
                                    backup_phone = await page.query_selector('button[aria-label^="الهاتف:"], button[aria-label^="Phone:"]')
                                    if backup_phone:
                                        p_text = await backup_phone.get_attribute('aria-label')
                                        phone = clean_phone(p_text.replace("الهاتف:", "").replace("Phone:", "").strip())

                            if website == "N/A":
                                web_button = await page.query_selector('a[data-item-id="authority"]')
                                if web_button: website = await web_button.get_attribute('href')
                                else:
                                    backup_web = await page.query_selector('a[aria-label*="Website:"], a[aria-label*="الموقع الإلكتروني:"], a[aria-label*="المواعيد:"]')
                                    if backup_web: website = await backup_web.get_attribute('href')

                    seen.add(name)
                    results.append({
                        "Business Name": name, "Maps URL": url, "Phone": phone,
                        "Rating": rating, "Reviews": reviews, "Website": website if website else "N/A"
                    })
                    print(f"[+] {len(results)}. {name[:30]}.. | Phone: {phone} | Rating: {rating} | Reviews: {reviews} | Web: {short_url(website) if website else 'N/A'}")
                except Exception as e:
                    pass
            
            last_cards_count = len(cards)
            feed = await page.query_selector('div[role="feed"]')
            if feed: await feed.evaluate("(el) => el.scrollBy(0, 3500)")
            await page.wait_for_timeout(2500)

        await browser.close()

        csv_path = os.path.join(OUTPUT_DIR, "cleaned_medical_leads.csv")
        excel_path = os.path.join(OUTPUT_DIR, "cleaned_medical_leads.xlsx")

        if results:
            df = pd.DataFrame(results)
            df.to_csv(csv_path, index=False, encoding="utf-8-sig")
            save_excel(results, excel_path)
            print(f"\n[+] Extraction successfully completed!")
        else:
            print("[!] No data collected.")

asyncio.run(scrape_google_maps("Dental Clinics in Dubai", total_results_needed=80))
